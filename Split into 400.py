import openpyxl
import os

def split_excel():
    desktop_path = os.path.join(os.path.expanduser('~'), 'Desktop')
    filename = os.path.join(desktop_path, 'AMA22032024.xlsx')
    
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    asset_ids = [cell.value for cell in ws['A'] if cell.value is not None]

    num_sheets = len(asset_ids) // 400 + (len(asset_ids) % 400 > 0)

    for i in range(num_sheets):
        new_ws = wb.create_sheet(title=f'AMA-{i+1}')
        start_row = i * 400
        end_row = min((i + 1) * 400, len(asset_ids))
        for j in range(start_row, end_row):
            new_ws.cell(row=j - start_row + 1, column=1, value=asset_ids[j])

    wb.remove(ws)
    wb.save(filename)

# Example usage
split_excel()
